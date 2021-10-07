def generate_marksheet():

    from openpyxl import Workbook
import openpyxl
import csv
import os

def fill_overall():
    
    with open("names-roll.csv",'r') as RollNames:
        rnRead=csv.reader(RollNames)
        studentDictionary={}
        for student in rnRead:
            studentDictionary[student[0]] = student[1]
    
    gradesDictionary={
        'AA':10,
        'AB':9,
        'BB':8,
        'BC':7,
        'CC':6,
        'CD':5,
        'DD':4,
        'F':0,
        'I':0
    }

    with open("grades.csv",'r') as grades:   #opening grade.csv file
        
        Graderead=csv.reader(grades) 
        cnt=0
        for readGra in Graderead:        
            if cnt ==0:
                cnt=1
                continue
            wBook =openpyxl.load_workbook(r'output/'+readGra[0]+'.xlsx')
            sheet=wBook['overall']
            
            # ******************* row 1 *********************************
            
            sheet.cell(row=1,column=1).value ='Roll No '
            sheet.cell(row=1,column=2).value = readGra[0]
            
            # ***************** row 2 ****************************
            
            sheet.cell(row=2,column=1).value ='Name of Student'
            sheet.cell(row=2,column=2).value = studentDictionary[str(readGra[0])]
           
            # ***************** row 3 ****************************
            
            sheet.cell(row=3,column=1).value ='Discipline'
            sheet.cell(row=3,column=2).value = readGra[0][4:6]
            
            ################ row 4 ###############
            
            sheet.cell(row=4,column=1).value ='Semester Number'
            for n in range(1,9):
                sheet.cell(row = 4, column = n+1).value = n
            
            ################ row 5 ###############
            sheet.cell(row=5,column=1).value ='Semester wise Credit taken'
            
            ################ row 6 ###############
            sheet.cell(row=6,column=1).value ='SPI'
            
            ################ row 7 ###############
            sheet.cell(row=7,column=1).value ='Total Credit Taken'
            
            ################ row 8 ###############
            sheet.cell(row=8,column=1).value ='CPI'

            sum=0
            temp=0
            temp2 = 0
            temp3 = 0 
            i=1
            # loop to fill Semester Wise Credit Taken
            while i < 9:
                src=wBook['sem'+str(i)]
                sum=0 
                row_inserted=src.max_row+1
                for rw in range(2,row_inserted):
                    n = src.cell(row=rw,column=5).value
                    sum =sum + n
                sheet.cell(row=5,column=i+1).value =sum
                i = i+1
            
            # loop to fill Total Credit Taken
            i=2
            while i < 10:
                n = sheet.cell(row=5,column=i).value
                temp=temp + n
                sheet.cell(row = 7, column = i).value = temp
                i = i+1   
            
            # loop to fill SPI
            i=1
            while i <= 8:
                sum=0
                src=wBook['sem'+str(i)]
                row_inserted=src.max_row
                for rw in range(2,row_inserted+1):
                    n = src.cell(row=rw,column=5).value * gradesDictionary[src.cell(row=rw,column=7).value]
                    sum = sum+n
                temp = sheet.cell(row=5,column=i+1).value
                if(temp==0):
                    sum = 1
                else:
                    sum = sum/temp 
                sheet.cell(row=6,column=i+1).value =sum
                i = i+1

            # loop to fill CPI
            i=2
            temp=0
            while i <= 9:
                n = sheet.cell(row=5,column=i).value * sheet.cell(row=6,column=i).value
                temp=temp+n
                temp2 = sheet.cell(row=7,column=i).value
                if(temp2 > 0 ):
                    temp3 = temp / temp2
                else:
                    temp3 = 1
                sheet.cell(row = 8, column = i).value = temp3     
                i = i+1
            wBook.save(r'output/'+readGra[0]+'.xlsx')
    return           

     
def genSs():
    with open("subjects_master.csv",'r') as subMaster:  # open subjects_master
        subMasterRead=csv.reader(subMaster)             # create row read of subjects_master
        subCodeNames={}                                 # define a dictionary
        for var in subMasterRead:                       # loop all the rows of subjects_master.csv
            subCodeNames[var[0]] = var[1], var[2]       # fill the dictionary
        
    with open("grades.csv",'r') as grades:              # open grades.csv
        readGrades=csv.reader(grades)                   # create row read of grades
        cnt = 0
        pam = 1
        for grd in readGrades:      # loop all the rows except first of grades.csv
            if pam > 200:
                break
            if cnt == 0:
                cnt = 1
                continue
            if os.path.exists(r'output/'+grd[0]+'.xlsx'):                   # check if our file exists
                wBook =openpyxl.load_workbook(r'output/'+grd[0]+'.xlsx')    # open the file as wBook
                i = 1
                while i < 9:                                     # make a loop from 1 to 8 inclusive
                    if(grd[1] == str(i)):                                 # if correct sem found for grd
                        tmp ='sem'+str(i)
                        Wsheet=wBook[tmp]                         # select the sheet
                        fRows=Wsheet.max_row+1                                 # find the number of rows filled in the sheet
                        Wsheet.cell(row=fRows,column=1).value = fRows      # Serial No
                        Wsheet.cell(row=fRows,column=2).value = grd[2]     # Subject Code
                        Wsheet.cell(row=fRows,column=3).value = subCodeNames[grd[2]][0]    # Subject Name
                        Wsheet.cell(row=fRows,column=4).value = subCodeNames[grd[2]][1]    # L-T-P
                        Wsheet.cell(row=fRows,column=5).value = int( grd[3] )              # course Credit
                        Wsheet.cell(row=fRows,column=6).value = grd[-1]                    # Course Type
                        Wsheet.cell(row=fRows,column=7).value = grd[-2]                    # Grade scored
                    i= i+1                                                               
            else:           # if the file do not exist create one
                wBook = openpyxl.Workbook() 
                overall = wBook.active
                wBook.create_sheet( index=0, title="overall")                           # create the front sheet as overall
                i = 1
                while i < 9:                                                            # loop [1,8]
                    wBook.create_sheet(index=i,title=r"sem"+str(i))                     # create various sheets
                    Wsheet=wBook['sem'+str(i)]                                           # select current semester sheet as according to loop
                    Wsheet.cell(row=1 ,column=1).value = 'Sl No'
                    Wsheet.cell(row=1 ,column=2).value = 'Subject No'
                    Wsheet.cell(row=1 ,column=3).value = 'Subject Name'
                    Wsheet.cell(row=1 ,column=4).value = 'L-T-P'
                    Wsheet.cell(row=1 ,column=5).value = 'Credit'
                    Wsheet.cell(row=1 ,column=6).value = 'Sub Type'
                    Wsheet.cell(row=1 ,column=7).value = 'Grade'
                    i= i+1
            pam = pam + 1
            wBook.save(r'output/'+grd[0]+'.xlsx')                             
    return

genSs()
fill_overall()


