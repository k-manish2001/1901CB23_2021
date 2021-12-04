from django.shortcuts import render,HttpResponse
from openpyxl.styles import colors
from home.models import Index
from openpyxl import Workbook
import pandas as pd
import openpyxl
import csv
import os
import fpdf
from fpdf import FPDF
from fpdf import FPDF_FONT_DIR 
from datetime import date
# Create your views here.
def generate_transcript(roll,f1,f2):
    class PDF(FPDF):
        def header(self):
            self.image('image/header.png',5,5,290)
            # self.image('image/logo.iitpatna.png',30,5,20)
            # self.image('image/logo.iitpatna.png',262,5,20)
            # self.line(5.0,30.0,292.0,30.0)
            self.rect(5.0, 5.0, 287.0,190.0)
            pdf.set_font('helvetica','B',26)
    
            # self.cell(0,2,"Indian Institute Of Technology Patna",border=False,ln=1,align='C')
            pdf.set_font('helvetica','B',16)
            
              

      
            
    pdf=PDF('p','mm','A3')
    pdf.add_page()
    pdf.line(5.0,80.0,292.0,80.0)
    pdf.line(5.0,120.0,292.0,120.0)
    # pdf.line(35.0,5.0,35.0,30.0)
    # pdf.line(260.0,5.0,260.0,30.0)
    with open("sample_input/names-roll.csv",'r') as R:
        r_read=csv.reader(R)
        roll_name={}
        for info in r_read:
            roll_name[info[0]]=info[1]
    pdf.set_font("Times", size=8)        
    pdf.set_xy(70.0,33.0)
    pdf.cell(15,0,"Roll No:",0,0) 
    pdf.set_xy(85.0,33.0)
    pdf.cell(15,0,str(roll),0,0)
    pdf.set_xy(110.0,33.0)
    pdf.cell(15,0,"Name:",0,0)
    pdf.set_xy(119.0,33.0)
    pdf.cell(30,0,str(roll_name[roll]),0,0)
    pdf.cell(35,0,"Year of Admission",0,0)
    pdf.set_xy(70.0,37.0)
    pdf.cell(15,0,"Programme:",0,0)
    pdf.cell(35,0,"Bechelor of Technology",0,0)
    pdf.cell(35,0,"Course:",0,0)
    pdf.rect(65.0, 31.0, 140.0,8.0)

          
    with open("sample_input/subjects_master.csv",'r') as NR:
        NR_read=csv.reader(NR)
        sub_name={}
        sub_ltp={} 
        for data in NR_read:
            sub_name[data[0]]=data[1]
            sub_ltp[data[0]]=data[2]
           
    with open("sample_input/grades.csv",'r') as name:
        
        name_read=csv.reader(name)

        sub1=["Sub Code"]
        sub2=["Sub Code"]
        sub3=["Sub Code"]
        sub4=["Sub Code"]
        sub5=["Sub Code"]
        sub6=["Sub Code"]
        sub7=["Sub Code"]
        sub8=["Sub Code"]
        name1=["Subject Name"]
        name2=["Subject Name"]
        name3=["Subject Name"]
        name4=["Subject Name"]
        name5=["Subject Name"]
        name6=["Subject Name"]
        name7=["Subject Name"]
        name8=["Subject Name"]
        lt1=["L-T-P"]
        lt2=["L-T-P"]
        lt3=["L-T-P"]
        lt4=["L-T-P"]
        lt5=["L-T-P"]
        lt6=["L-T-P"]
        lt7=["L-T-P"]
        lt8=["L-T-P"]
        cr1=["CRD"]
        cr2=["CRD"]
        cr2=["CRD"]
        cr2=["CRD"]
        cr3=["CRD"]
        cr4=["CRD"]
        cr5=["CRD"]
        cr6=["CRD"]
        cr7=["CRD"]
        cr8=["CRD"]
        gr2=["GRD"]
        gr1=["GRD"]
        gr2=["GRD"]
        gr2=["GRD"]
        gr2=["GRD"]
        gr3=["GRD"]
        gr4=["GRD"]
        gr5=["GRD"]
        gr6=["GRD"]
        gr7=["GRD"]
        gr8=["GRD"]
        t=0   
        check=0
        for info in name_read:
            if roll==info[0]:
                check+=1 
                break
        if check==0:
            return    
        for info in name_read:
            
            if info[0]!=roll:
                continue
            if info[1]=="1" and info[0]==roll: 
                sub1.append(info[2])
                name1.append(sub_name[info[2]])
                lt1.append(sub_ltp[info[2]])
                cr1.append(info[3])
                gr1.append(info[4])
            if info[1]=="2" and info[0]==roll:
                sub2.append(info[2])
                name2.append(sub_name[info[2]])
                lt2.append(sub_ltp[info[2]])
                cr2.append(info[3])
                gr2.append(info[4])
            if info[1]=="3" and info[0]==roll:
                sub3.append(info[2])
                name3.append(sub_name[info[2]])
                lt3.append(sub_ltp[info[2]])
                cr3.append(info[3])
                gr3.append(info[4])
            if info[1]=="4" and info[0]==roll:
                sub4.append(info[2])
                name4.append(sub_name[info[2]])
                lt4.append(sub_ltp[info[2]])
                cr4.append(info[3])
                gr4.append(info[4]) 
            if info[1]=="5" and info[0]==roll:
                sub5.append(info[2])
                name5.append(sub_name[info[2]])
                lt5.append(sub_ltp[info[2]])
                cr5.append(info[3])
                gr5.append(info[4])
            if info[1]=="6" and info[0]==roll:
                sub6.append(info[2])
                name6.append(sub_name[info[2]])
                lt6.append(sub_ltp[info[2]])
                cr6.append(info[3])
                gr6.append(info[4])
            if info[1]=="7" and info[0]==roll:
                sub7.append(info[2])
                name7.append(sub_name[info[2]])
                lt7.append(sub_ltp[info[2]])
                cr7.append(info[3])
                gr7.append(info[4])                

            if info[1]=="8" and info[0]==roll:
                sub8.append(info[2])
                name8.append(sub_name[info[2]])
                lt8.append(sub_ltp[info[2]])
                cr8.append(info[3])
                gr8.append(info[4])
        grade_dic= { 'AA':10,'AA*':10,'AB':9,'AB*':9,'BB':8,'BB*':8,'BC':7,'BC*':7,'CC':6,'CC*':6,'CD':5,'CD*':5,'DD':4,'DD*':4,'F':0,'F*':0,'I':0,'I*':0}
        if len(sub1)>0:           
            data_dict={"Sub Code":sub1,
                      "Sub Name": name1,
                       "L-T-P":lt1,
                       "CRD":cr1,
                       "GRD":gr1
                       }
            pdf.set_font("Times", size=5)
            pdf.set_xy(5.0,36.0)
            pdf.cell(1,10,"SEM1",border=False,ln=1)
            line_height = pdf.font_size * 2.5
            col_width = pdf.epw / 4  # distribute content evenly
            pdf.line(6.0,42.0,100.0,42.0)
            t=44
            crd_taken1=0
            grd_clear=0
            total=0
            for j in range(len(data_dict["Sub Code"])):
                if j==0:
                    continue
                else:
                    crd_taken1+=int(data_dict["CRD"][j])
                    total+=int(data_dict["CRD"][j])*grade_dic[data_dict["GRD"][j]]
                    if grade_dic[data_dict["GRD"][j]]>0:
                        grd_clear+=int(data_dict["CRD"][j])
            sem1_spi=total/crd_taken1
            sem1_cpi=sem1_spi


            for i in range(len(data_dict["Sub Code"])):
                pdf.set_xy(5.0,44.0+(i*3.0))
                t+=3
                pdf.line(6.0,45.0+(i*3.0),100.0,45.0+(i*3.0))
                pdf.line(6.0,45.0+(i*3.0),6.0,42.0+(i*3.0))
                pdf.line(100.0,45.0+(i*3.0),100.0,42.0+(i*3.0))
                pdf.line(15.0,45.0+(i*3.0),15.0,42.0+(i*3.0))
                pdf.line(60.0,45.0+(i*3.0),60.0,42.0+(i*3.0))
                pdf.line(75.0,45.0+(i*3.0),75.0,42.0+(i*3.0))
                pdf.line(90.0,45.0+(i*3.0),90.0,42.0+(i*3.0))
                pdf.multi_cell(15,0, data_dict["Sub Code"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(45,0, data_dict["Sub Name"][i],0,0,max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["L-T-P"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["CRD"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["GRD"][i],0,0, max_line_height=pdf.font_size)
                
                pdf.ln()
            pdf.line(6.0,t,100.0,t) 
            pdf.line(6.0,t+5,100.0,t+5)
            pdf.line(6.0,t+5,6.0,t) 
            pdf.line(100.0,t+5,100.0,t) 
            pdf.set_xy(7,t-2) 
            pdf.set_font("Times", size=5)                
            pdf.cell(15,10,"Credits Taken:",0,0)
            pdf.set_xy(18,t+0.5) 
            pdf.cell(15,5,str(crd_taken1),0,0)
            pdf.set_xy(21,t+3)
            pdf.cell(15,0,"Credits Cleared",0,0)
            pdf.set_xy(33,t+3)
            pdf.cell(15,0,str(grd_clear),0,0)
            pdf.set_xy(37,t+3)
            pdf.cell(15,0,"SPI",0,0)
            pdf.set_xy(40,t+3)
            pdf.cell(15,0,str(round(sem1_spi,2)),0,0)
            pdf.set_xy(47,t+3)
            pdf.cell(15,0,"CPI",0,0)
            pdf.set_xy(50,t+3)
            pdf.cell(15,0,str(round(sem1_cpi,2)),0,0)       
        if len(sub2)>0:
            data_dict={"Sub Code":sub2,
                       "Sub Name":name2,
                       "L-T-P":lt2,
                       "CRD":cr2,
                       "GRD":gr2
                      }
            crd_taken2=0
            grd_clear=0
            total=0
            for j in range(len(data_dict["Sub Code"])):
                if j==0:
                    continue
                else:
                    crd_taken2+=int(data_dict["CRD"][j])
                    total+=int(data_dict["CRD"][j])*grade_dic[data_dict["GRD"][j]]
                    if grade_dic[data_dict["GRD"][j]]>0:
                        grd_clear+=int(data_dict["CRD"][j])
            sem2_spi=total/crd_taken2
            sem2_cpi=((sem1_spi*crd_taken1)+(sem2_spi*crd_taken2))/(crd_taken2+crd_taken1)            
            pdf.set_font("Times", size=5)
            line_height = pdf.font_size * 2.5
            pdf.set_xy(110.0,36.0)
            pdf.cell(1,10,"SEM2",border=False,ln=1)
            col_width = pdf.epw / 4       # distribute content evenly
            pdf.line(111.0,42.0,199.0,42.0)
            t=43
            for i in range(len(data_dict["Sub Code"])):
                    t+=3
                    pdf.set_xy(110.0,44.0+(i*3.0))
                    pdf.line(111.0,45.0+(i*3.0),199.0,45.0+(i*3.0))
                    pdf.line(111.0,45.0+(i*3.0),111.0,42.0+(i*3.0))
                    pdf.line(199.0,45.0+(i*3.0),199.0,42.0+(i*3.0))
                    pdf.line(120.0,45.0+(i*3.0),120.0,42.0+(i*3.0))
                    pdf.line(165.0,45.0+(i*3.0),165.0,42.0+(i*3.0))
                    pdf.line(180.0,45.0+(i*3.0),180.0,42.0+(i*3.0))
                    pdf.line(190.0,45.0+(i*3.0),190.0,42.0+(i*3.0))                    
                    pdf.multi_cell(15,0, data_dict["Sub Code"][i],0,0, max_line_height=pdf.font_size)
                    pdf.multi_cell(40,0, data_dict["Sub Name"][i],0,0,max_line_height=pdf.font_size)
                    pdf.multi_cell(15,0, data_dict["L-T-P"][i],0,0, max_line_height=pdf.font_size)
                    pdf.multi_cell(15,0, data_dict["CRD"][i],0,0, max_line_height=pdf.font_size)
                    pdf.multi_cell(15,0, data_dict["GRD"][i],0,0, max_line_height=pdf.font_size)
                    pdf.ln() 
            pdf.line(111.0,t+1,199.0,t+1) 
            pdf.line(111.0,t+6,199.0,t+6)
            pdf.line(111.0,t+6,111.0,t+1) 
            pdf.line(199.0,t+6,199.0,t+1)  
            pdf.set_font("Times", size=5)
            pdf.set_xy(111,t-1) 
            pdf.cell(15,10,"Credits Taken:",0,0)
            pdf.set_xy(124,t+1.5) 
            pdf.cell(15,5,str(crd_taken2),0,0)
            pdf.set_xy(127,t+4)
            pdf.cell(15,0,"Credits Cleared",0,0)
            pdf.set_xy(139,t+4)
            pdf.cell(15,0,str(grd_clear),0,0)
            pdf.set_xy(143,t+4)
            pdf.cell(15,0,"SPI",0,0)
            pdf.set_xy(146,t+4)
            pdf.cell(15,0,str(round(sem2_spi,2)),0,0)
            pdf.set_xy(153,t+4)
            pdf.cell(15,0,"CPI",0,0)
            pdf.set_xy(156,t+4)
            pdf.cell(15,0,str(round(sem2_cpi,2)),0,0)                                                       
        if len(sub3)>0:
            data_dict={"Sub Code":sub3,
                       "Sub Name":name3,
                       "L-T-P":lt3,
                       "CRD":cr3,
                       "GRD":gr3
                      } 
            crd_taken3=0
            grd_clear=0
            total=0
            for j in range(len(data_dict["Sub Code"])):
                if j==0:
                    continue
                else:
                    crd_taken3+=int(data_dict["CRD"][j])
                    total+=int(data_dict["CRD"][j])*grade_dic[data_dict["GRD"][j]]
                    if grade_dic[data_dict["GRD"][j]]>0:
                        grd_clear+=int(data_dict["CRD"][j]) 
            sem3_spi=total/crd_taken3
            sem3_cpi=((sem1_spi*crd_taken1)+(sem2_spi*crd_taken2)+(sem3_spi*crd_taken3))/(crd_taken2+crd_taken3+crd_taken1)                                               
            pdf.set_font("Times", size=5)
            line_height = pdf.font_size * 2.5
            pdf.set_xy(200.0,36.0)
            pdf.cell(1,10,"SEM3",border=False,ln=1)
            col_width = pdf.epw / 4       # distribute content evenly
            pdf.line(201.0,42.0,290.0,42.0)
            t=43
            for i in range(len(data_dict["Sub Code"])):
                    pdf.set_xy(200.0,44.0+(i*3.0))
                    t+=3
                    pdf.line(201.0,45.0+(i*3.0),290.0,45.0+(i*3.0))
                    pdf.line(201.0,45.0+(i*3.0),201.0,42.0+(i*3.0))
                    pdf.line(290.0,45.0+(i*3.0),290.0,42.0+(i*3.0))
                    pdf.line(210.0,45.0+(i*3.0),210.0,42.0+(i*3.0))
                    pdf.line(255.0,45.0+(i*3.0),255.0,42.0+(i*3.0))
                    pdf.line(270.0,45.0+(i*3.0),270.0,42.0+(i*3.0))
                    pdf.line(285.0,45.0+(i*3.0),285.0,42.0+(i*3.0))                    
                    pdf.multi_cell(15,0, data_dict["Sub Code"][i],0,0, max_line_height=pdf.font_size)
                    pdf.multi_cell(40,0, data_dict["Sub Name"][i],0,0,max_line_height=pdf.font_size)
                    pdf.multi_cell(15,0, data_dict["L-T-P"][i],0,0, max_line_height=pdf.font_size)
                    pdf.multi_cell(15,0, data_dict["CRD"][i],0,0, max_line_height=pdf.font_size)
                    pdf.multi_cell(15,0, data_dict["GRD"][i],0,0, max_line_height=pdf.font_size)
                    pdf.ln(line_height)
            pdf.line(201.0,t+1,285.0,t+1) 
            pdf.line(201.0,t+6,285.0,t+6)
            pdf.line(201.0,t+6,201.0,t+1) 
            pdf.line(285.0,t+6,285.0,t+1)  
            pdf.set_font("Times", size=5)
            pdf.set_xy(201,t-1) 
            pdf.cell(15,10,"Credits Taken:",0,0)
            pdf.set_xy(214,t+1.5) 
            pdf.cell(15,5,str(crd_taken3),0,0)
            pdf.set_xy(227,t+4)
            pdf.cell(15,0,"Credits Cleared",0,0)
            pdf.set_xy(239,t+4)
            pdf.cell(15,0,str(grd_clear),0,0)
            pdf.set_xy(243,t+4)
            pdf.cell(15,0,"SPI",0,0)
            pdf.set_xy(246,t+4)
            pdf.cell(15,0,str(round(sem3_spi,2)),0,0)
            pdf.set_xy(253,t+4)
            pdf.cell(15,0,"CPI",0,0)
            pdf.set_xy(256,t+4)
            pdf.cell(15,0,str(round(sem3_cpi,2)),0,0)                                           
        if len(sub4)>0:
            data_dict={"Sub Code":sub4,
                       "Sub Name":name4,
                       "L-T-P":lt4,
                       "CRD":cr4,
                       "GRD":gr4
                      } 
            crd_taken4=0
            grd_clear=0
            total=0                      
            for j in range(len(data_dict["Sub Code"])):
                if j==0:
                    continue
                else:
                    crd_taken4+=int(data_dict["CRD"][j])
                    total+=int(data_dict["CRD"][j])*grade_dic[data_dict["GRD"][j]]
                    if grade_dic[data_dict["GRD"][j]]>0:
                        grd_clear+=int(data_dict["CRD"][j]) 
            sem4_spi=total/crd_taken4
            sem4_cpi=((sem1_spi*crd_taken1)+(sem2_spi*crd_taken2)+(sem3_spi*crd_taken3)+(sem4_spi*crd_taken4))/(crd_taken2+crd_taken3+crd_taken1+crd_taken4)                      
            pdf.set_font("Times", size=5)
            pdf.set_xy(5.0,78.0)
            pdf.cell(1,10,"SEM4",border=False,ln=1)
            line_height = pdf.font_size * 2.5
            col_width = pdf.epw / 4  # distribute content evenly
            pdf.line(6.0,84.0,100.0,84.0)
            t=86
            for i in range(len(data_dict["Sub Code"])):
                pdf.set_xy(5.0,86.0+(i*3.0))
                t+=3
                pdf.line(6.0,87.0+(i*3.0),100.0,87.0+(i*3.0))
                pdf.line(6.0,87.0+(i*3.0),6.0,84.0+(i*3.0))
                pdf.line(100.0,87.0+(i*3.0),100.0,84.0+(i*3.0))
                pdf.line(15.0,87.0+(i*3.0),15.0,84.0+(i*3.0))
                pdf.line(60.0,87.0+(i*3.0),60.0,84.0+(i*3.0))
                pdf.line(75.0,87.0+(i*3.0),75.0,84.0+(i*3.0))
                pdf.line(90.0,87.0+(i*3.0),90.0,84.0+(i*3.0))
                pdf.multi_cell(15,0, data_dict["Sub Code"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(45,0, data_dict["Sub Name"][i],0,0,max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["L-T-P"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["CRD"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["GRD"][i],0,0, max_line_height=pdf.font_size)
                
                pdf.ln()
            pdf.line(6.0,t,100.0,t) 
            pdf.line(6.0,t+5,100.0,t+5)
            pdf.line(6.0,t+5,6.0,t) 
            pdf.line(100.0,t+5,100.0,t)  
            pdf.set_font("Times", size=5)
            pdf.set_xy(6,t-1) 
            pdf.cell(15,10,"Credits Taken:",0,0)
            pdf.set_xy(19,t+1.5) 
            pdf.cell(15,5,str(crd_taken4),0,0)
            pdf.set_xy(24,t+4)
            pdf.cell(15,0,"Credits Cleared",0,0)
            pdf.set_xy(36,t+4)
            pdf.cell(15,0,str(grd_clear),0,0)
            pdf.set_xy(44,t+4)
            pdf.cell(15,0,"SPI",0,0)
            pdf.set_xy(50,t+4)
            pdf.cell(15,0,str(round(sem4_spi,2)),0,0)
            pdf.set_xy(57,t+4)
            pdf.cell(15,0,"CPI",0,0)
            pdf.set_xy(60,t+4)
            pdf.cell(15,0,str(round(sem4_cpi,2)),0,0)                                    
        if len(sub5)>0:
            data_dict={"Sub Code":sub5,
                       "Sub Name":name5,
                       "L-T-P":lt5,
                       "CRD":cr5,
                       "GRD":gr5
                      }   
            crd_taken5=0
            grd_clear=0
            total=0                      
            for j in range(len(data_dict["Sub Code"])):
                if j==0:
                    continue
                else:
                    crd_taken5+=int(data_dict["CRD"][j])
                    total+=int(data_dict["CRD"][j])*grade_dic[data_dict["GRD"][j]]
                    if grade_dic[data_dict["GRD"][j]]>0:
                        grd_clear+=int(data_dict["CRD"][j]) 
            sem5_spi=total/crd_taken5
            sem5_cpi=((sem1_spi*crd_taken1)+(sem2_spi*crd_taken2)+(sem3_spi*crd_taken3)+(sem4_spi*crd_taken4)+(sem5_spi*crd_taken5))/(crd_taken2+crd_taken3+crd_taken1+crd_taken4+crd_taken5)        
            pdf.set_font("Times", size=5)
            pdf.set_xy(102.0,78.0)
            pdf.cell(1,10,"SEM5",border=False,ln=1)
            line_height = pdf.font_size * 2.5
            col_width = pdf.epw / 4  # distribute content evenly
            pdf.line(102.0,84.0,196.0,84.0)
            t=86
            for i in range(len(data_dict["Sub Code"])):
                pdf.set_xy(102.0,86.0+(i*3.0))
                t+=3
                pdf.line(102.0,87.0+(i*3.0),196.0,87.0+(i*3.0))
                pdf.line(102.0,87.0+(i*3.0),102.0,84.0+(i*3.0))
                pdf.line(196.0,87.0+(i*3.0),196.0,84.0+(i*3.0))
                pdf.line(111.0,87.0+(i*3.0),111.0,84.0+(i*3.0))
                pdf.line(156.0,87.0+(i*3.0),156.0,84.0+(i*3.0))
                pdf.line(171.0,87.0+(i*3.0),171.0,84.0+(i*3.0))
                pdf.line(186.0,87.0+(i*3.0),186.0,84.0+(i*3.0))
                pdf.multi_cell(15,0, data_dict["Sub Code"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(45,0, data_dict["Sub Name"][i],0,0,max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["L-T-P"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["CRD"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["GRD"][i],0,0, max_line_height=pdf.font_size)
                
                pdf.ln()
            pdf.line(102.0,t,196.0,t) 
            pdf.line(102.0,t+5,196.0,t+5)
            pdf.line(102.0,t+5,102.0,t) 
            pdf.line(196.0,t+5,196.0,t)  
            pdf.set_font("Times", size=5)
            pdf.set_xy(105,t-1) 
            pdf.cell(15,10,"Credits Taken:",0,0)
            pdf.set_xy(118,t+1.5) 
            pdf.cell(15,5,str(crd_taken5),0,0)
            pdf.set_xy(123,t+4)
            pdf.cell(15,0,"Credits Cleared:",0,0)
            pdf.set_xy(137,t+4)
            pdf.cell(15,0,str(grd_clear),0,0)
            pdf.set_xy(146,t+4)
            pdf.cell(15,0,"SPI:",0,0)
            pdf.set_xy(151,t+4)
            pdf.cell(15,0,str(round(sem5_spi,2)),0,0)
            pdf.set_xy(164,t+4)
            pdf.cell(15,0,"CPI:",0,0)
            pdf.set_xy(169,t+4)
            pdf.cell(15,0,str(round(sem5_cpi,2)),0,0)            
                       

              
        if len(sub6)>0:
            data_dict={"Sub Code":sub5,
                       "Sub Name":name5,
                       "L-T-P":lt5,
                       "CRD":cr5,
                       "GRD":gr5
                      }   
            crd_taken6=0
            grd_clear=0
            total=0                      
            for j in range(len(data_dict["Sub Code"])):
                if j==0:
                    continue
                else:
                    crd_taken6+=int(data_dict["CRD"][j])
                    total+=int(data_dict["CRD"][j])*grade_dic[data_dict["GRD"][j]]
                    if grade_dic[data_dict["GRD"][j]]>0:
                        grd_clear+=int(data_dict["CRD"][j]) 
            sem6_spi=total/crd_taken6
            sem6_cpi=((sem1_spi*crd_taken1)+(sem2_spi*crd_taken2)+(sem3_spi*crd_taken3)+(sem4_spi*crd_taken4)+(sem5_spi*crd_taken5)+(sem6_spi*crd_taken6))/(crd_taken2+crd_taken3+crd_taken1+crd_taken4+crd_taken5+crd_taken6)        
            pdf.set_font("Times", size=5)
            pdf.set_xy(198.0,78.0)
            pdf.cell(1,10,"SEM6",border=False,ln=1)
            line_height = pdf.font_size * 2.5
            col_width = pdf.epw / 4  # distribute content evenly
            pdf.line(198.0,84.0,291.0,84.0)
            t=86
            for i in range(len(data_dict["Sub Code"])):
                pdf.set_xy(197.0,86.0+(i*3.0))
                t+=3
                pdf.line(198.0,87.0+(i*3.0),291.0,87.0+(i*3.0))
                pdf.line(198.0,87.0+(i*3.0),198.0,84.0+(i*3.0))
                pdf.line(291.0,87.0+(i*3.0),291.0,84.0+(i*3.0))
                pdf.line(207.0,87.0+(i*3.0),207.0,84.0+(i*3.0))
                pdf.line(252.0,87.0+(i*3.0),252.0,84.0+(i*3.0))
                pdf.line(267.0,87.0+(i*3.0),267.0,84.0+(i*3.0))
                pdf.line(182.0,87.0+(i*3.0),182.0,84.0+(i*3.0))
                pdf.multi_cell(15,0, data_dict["Sub Code"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(45,0, data_dict["Sub Name"][i],0,0,max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["L-T-P"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["CRD"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["GRD"][i],0,0, max_line_height=pdf.font_size)
                
                pdf.ln()
            pdf.line(198.0,t,291.0,t) 
            pdf.line(198.0,t+5,291.0,t+5)
            pdf.line(198.0,t+5,198.0,t) 
            pdf.line(291.0,t+5,291.0,t)  
            pdf.set_font("Times", size=5)
            pdf.set_xy(198,t-1) 
            pdf.cell(15,10,"Credits Taken:",0,0)
            pdf.set_xy(211,t+1.5) 
            pdf.cell(15,5,str(crd_taken6),0,0)
            pdf.set_xy(216,t+4)
            pdf.cell(15,0,"Credits Cleared:",0,0)
            pdf.set_xy(230,t+4)
            pdf.cell(15,0,str(grd_clear),0,0)
            pdf.set_xy(239,t+4)
            pdf.cell(15,0,"SPI:",0,0)
            pdf.set_xy(243,t+4)
            pdf.cell(15,0,str(round(sem6_spi,2)),0,0)
            pdf.set_xy(256,t+4)
            pdf.cell(15,0,"CPI:",0,0)
            pdf.set_xy(261,t+4)
            pdf.cell(15,0,str(round(sem6_cpi,2)),0,0)                                    
        if len(sub7)>0:
            data_dict={"Sub Code":sub8,
                       "Sub Name":name8,
                       "L-T-P":lt8,
                       "CRD":cr8,
                       "GRD":gr8
                      }
            crd_taken7=0
            grd_clear=0
            total=0                      
            for j in range(len(data_dict["Sub Code"])):
                if j==0:
                    continue
                else:
                    crd_taken7+=int(data_dict["CRD"][j])
                    total+=int(data_dict["CRD"][j])*grade_dic[data_dict["GRD"][j]]
                    if grade_dic[data_dict["GRD"][j]]>0:
                        grd_clear+=int(data_dict["CRD"][j]) 
            sem7_spi=total/crd_taken7
            sem7_cpi=((sem1_spi*crd_taken1)+(sem2_spi*crd_taken2)+(sem3_spi*crd_taken3)+(sem4_spi*crd_taken4)+(sem5_spi*crd_taken5)+(sem6_spi*crd_taken6)+(sem7_spi*crd_taken7))/(crd_taken2+crd_taken3+crd_taken1+crd_taken4+crd_taken5+crd_taken6+crd_taken7)
            pdf.set_font("Times", size=5)
            pdf.set_xy(5.0,123.0)
            pdf.cell(1,10,"SEM7",border=False,ln=1)
            line_height = pdf.font_size * 2.5
            col_width = pdf.epw / 4  # distribute content evenly
            pdf.line(6.0,129.0,100.0,129.0)
            t=132
            for i in range(len(data_dict["Sub Code"])):
                pdf.set_xy(5.0,130.0+(i*3.0))
                t+=3
                pdf.line(6.0,132.0+(i*3.0),100.0,132.0+(i*3.0))
                pdf.line(6.0,132.0+(i*3.0),6.0,129.0+(i*3.0))
                pdf.line(100.0,132.0+(i*3.0),100.0,129.0+(i*3.0))
                pdf.line(15.0,132.0+(i*3.0),15.0,129.0+(i*3.0))
                pdf.line(60.0,132.0+(i*3.0),60.0,129.0+(i*3.0))
                pdf.line(75.0,132.0+(i*3.0),75.0,129.0+(i*3.0))
                pdf.line(90.0,132.0+(i*3.0),90.0,129.0+(i*3.0))
                pdf.multi_cell(15,0, data_dict["Sub Code"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(45,0, data_dict["Sub Name"][i],0,0,max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["L-T-P"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["CRD"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["GRD"][i],0,0, max_line_height=pdf.font_size)
                
                pdf.ln()
            pdf.line(6.0,t,100.0,t) 
            pdf.line(6.0,t+5,100.0,t+5)
            pdf.line(6.0,t+5,6.0,t) 
            pdf.line(100.0,t+5,100.0,t)  
            pdf.set_font("Times", size=5)
            pdf.set_xy(6,t-1) 
            pdf.cell(15,10,"Credits Taken:",0,0)
            pdf.set_xy(19,t+1.5) 
            pdf.cell(15,5,str(crd_taken4),0,0)
            pdf.set_xy(24,t+4)
            pdf.cell(15,0,"Credits Cleared",0,0)
            pdf.set_xy(36,t+4)
            pdf.cell(15,0,str(grd_clear),0,0)
            pdf.set_xy(44,t+4)
            pdf.cell(15,0,"SPI",0,0)
            pdf.set_xy(50,t+4)
            pdf.cell(15,0,str(round(sem7_spi,2)),0,0)
            pdf.set_xy(57,t+4)
            pdf.cell(15,0,"CPI",0,0)
            pdf.set_xy(60,t+4)
            pdf.cell(15,0,str(round(sem7_cpi,2)),0,0)              
        if len(sub8)>0:
            data_dict={"Sub Code":sub8,
                       "Sub Name":name8,
                       "L-T-P":lt8,
                       "CRD":cr8,
                       "GRD":gr8
                      }
            crd_taken8=0
            grd_clear=0
            total=0                      
            for j in range(len(data_dict["Sub Code"])):
                if j==0:
                    continue
                else:
                    crd_taken8+=int(data_dict["CRD"][j])
                    total+=int(data_dict["CRD"][j])*grade_dic[data_dict["GRD"][j]]
                    if grade_dic[data_dict["GRD"][j]]>0:
                        grd_clear+=int(data_dict["CRD"][j]) 
            sem8_spi=total/crd_taken8
            sem8_cpi=((sem1_spi*crd_taken1)+(sem2_spi*crd_taken2)+(sem3_spi*crd_taken3)+(sem4_spi*crd_taken4)+(sem5_spi*crd_taken5)+(sem6_spi*crd_taken6)+(sem7_spi*crd_taken7)+(sem8_spi*crd_taken8))/(crd_taken2+crd_taken3+crd_taken1+crd_taken4+crd_taken5+crd_taken6+crd_taken7+crd_taken8)                      
            pdf.set_font("Times", size=5)
            pdf.set_xy(102.0,123.0)
            pdf.cell(1,10,"SEM8",border=False,ln=1)
            line_height = pdf.font_size*2.5
            col_width = pdf.epw / 4  # distribute content evenly
            pdf.line(102.0,129.0,198.0,129.0)
            t=132
            for i in range(len(data_dict["Sub Code"])):
                pdf.set_xy(102.0,130.0+(i*3.0))
                t+=3
                pdf.line(102.0,132.0+(i*3.0),198.0,132.0+(i*3.0))
                pdf.line(102.0,132.0+(i*3.0),102.0,129.0+(i*3.0))
                pdf.line(198.0,132.0+(i*3.0),198.0,129.0+(i*3.0))
                pdf.line(111.0,132.0+(i*3.0),111.0,129.0+(i*3.0))
                pdf.line(156.0,132.0+(i*3.0),156.0,129.0+(i*3.0))
                pdf.line(171.0,132.0+(i*3.0),171.0,129.0+(i*3.0))
                pdf.line(186.0,132.0+(i*3.0),186.0,129.0+(i*3.0))
                pdf.multi_cell(15,0, data_dict["Sub Code"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(45,0, data_dict["Sub Name"][i],0,0,max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["L-T-P"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["CRD"][i],0,0, max_line_height=pdf.font_size)
                pdf.multi_cell(15,0, data_dict["GRD"][i],0,0, max_line_height=pdf.font_size)
                
                pdf.ln()
            pdf.line(102.0,t,198.0,t) 
            pdf.line(102.0,t+5,198.0,t+5)
            pdf.line(102.0,t+5,102.0,t) 
            pdf.line(198.0,t+5,198.0,t)  
            pdf.set_font("Times", size=5)
            pdf.set_xy(105,t-1) 
            pdf.cell(15,10,"Credits Taken:",0,0)
            pdf.set_xy(118,t+1.5) 
            pdf.cell(15,5,str(crd_taken8),0,0)
            pdf.set_xy(123,t+4)
            pdf.cell(15,0,"Credits Cleared:",0,0)
            pdf.set_xy(137,t+4)
            pdf.cell(15,0,str(grd_clear),0,0)
            pdf.set_xy(146,t+4)
            pdf.cell(15,0,"SPI:",0,0)
            pdf.set_xy(151,t+4)
            pdf.cell(15,0,str(round(sem8_spi,2)),0,0)
            pdf.set_xy(164,t+4)
            pdf.cell(15,0,"CPI:",0,0)
            pdf.set_xy(169,t+4)
            pdf.cell(15,0,str(round(sem8_cpi,2)),0,0)
    pdf.set_font("Times", size=10)
    pdf.set_xy(7.0,180) 
    pdf.cell(15,0,"Date of Issue",0,0)
    pdf.set_xy(29.0,178)
    pdf.cell(15,0,str(date.today()),0,0)

    pdf.line(29,181,51,181) 
    pdf.set_xy(230.0,180)
    pdf.cell(15,0,"Assitant Registrar(Academic)",0,0)
    pdf.line(230,178,270,178)
    if f1!=0:
        pdf.image(f1,95,160,20)
        pdf.image(f2,230,155,20)



    pdf.output("transcriptsIITP/"+roll+'.pdf')
    return    
def index(request):
    if request.method=='POST' :
        c=1
        try:
            n=request.POST.get('tem1')
            f1=request.FILES["file1"]
        

            f2=request.FILES["file2"]
            c+=1
        except:
            n=request.POST.get('tem1')
            
        x=n.split("-")
        roll_list=[]
        t=x[0]
        fir=int(x[0][-2]+x[0][-1])
        sec=int(x[1][-2]+x[1][-1])
        roll_list.append(t.upper())
        i=1
        while fir!=sec:
            if int(t[-2]+t[-1])<9:
                t=x[0][0:7]+str(int(x[0][-1])+i)
            else:
                t=x[0][0:6]+str(int(x[0][-2]+x[0][-1])+i)
            roll_list.append(t.upper())
            fir+=1
            i+=1
        for it in roll_list:
            try:
                generate_transcript(it,f1,f2)
            except:
                generate_transcript(it,0,0) 
        return HttpResponse("Your transcript Has Been Generated")          



              
                                    
    return render(request,'index.html')