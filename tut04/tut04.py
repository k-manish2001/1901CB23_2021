import openpyxl
from openpyxl import Workbook
import csv
#Manish kushwah
#1901CB23
#Driver function
import os

#output by roll numbers*************************
def output_individual_roll():
#opening the file
    open_file= open('regtable_old.csv','r')                
    with open_file:   
        for Row in open_file:
            if os.path.exists(f"output_individual_roll/{Row[0]}.xlsx"):
                workbook=openpyxl.load_workbook(r'output_individual_roll/'+Row[0]+'.xlsx') 
                active_sheet=workbook.active

                sheet_present=active_sheet.max_row #counting no. of sheets
                #adding the data with same roll no.
                active_sheet.cell(row=sheet_present+1 ,column=1).value = Row[0] #for roll no.
                active_sheet.cell(row=sheet_present+1 ,column=2).value = Row[1] #for register_sem
                active_sheet.cell(row=sheet_present+1 ,column=3).value = Row[3] #for sub_no
                active_sheet.cell(row=sheet_present+1 ,column=4).value = Row[8] #for sub_type
            else: #if the file doesn't exist
                workbook = Workbook()
                active_sheet = workbook.active
                active_sheet.cell(row=1 ,column=1).value = 'Rollno'
                active_sheet.cell(row=1 ,column=2).value = 'Register_sem'
                active_sheet.cell(row=1 ,column=3).value = 'Subno'
                active_sheet.cell(row=1 ,column=4).value = 'Sub_type'
            workbook.save(r'output_individual_roll/'+Row[0]+'.xlsx')

    return



output_individual_roll()



#output by subjects*************************
def output_by_subject():
#opening the file
    open_file= open('regtable_old.csv','r')                
    with open_file:   
        for Row in open_file:
            if os.path.exists(f"output_individual_roll/{Row[3]}.xlsx"):
                workbook=openpyxl.load_workbook(r'output_individual_roll/'+Row[3]+'.xlsx') 
                active_sheet=workbook.active

                sheet_present=active_sheet.max_row #counting no. of sheets
                #adding the data with same roll no.
                active_sheet.cell(row=sheet_present+1 ,column=1).value = Row[0] #for roll no.
                active_sheet.cell(row=sheet_present+1 ,column=2).value = Row[1] #for register_sem
                active_sheet.cell(row=sheet_present+1 ,column=3).value = Row[3] #for sub_no
                active_sheet.cell(row=sheet_present+1 ,column=4).value = Row[8] #for sub_type
            else: #if the file doesn't exist
                workbook = Workbook()
                active_sheet = workbook.active
                active_sheet.cell(row=1 ,column=1).value = 'Rollno'
                active_sheet.cell(row=1 ,column=2).value = 'Register_sem'
                active_sheet.cell(row=1 ,column=3).value = 'Subno'
                active_sheet.cell(row=1 ,column=4).value = 'Sub_type'
            workbook.save(r'output_individual_roll/'+Row[3]+'.xlsx')

    return



output_by_subject()