import csv
import openpyxl
from openpyxl import Workbook

def feedback_not_submitted():

    
    ltp_mapping_feedback_type = {1: 'Lecture', 2: 'tutorial', 3:'practical'}
    output_file_name = "course_feedback_remaining.xlsx" 
    check=-1
    with open('course_registered_by_all_students.csv','r') as Registered_students_file:
          Read1=csv.reader(Registered_students_file)
          W=Workbook()

		  # Driver program
          sheet=W.active
          sheet.cell(row=1,column=8).value='contact'
          sheet.cell(row=1,column=7).value='aemail'
          sheet.cell(row=1,column=6).value='email'
          sheet.cell(row=1,column=5).value='Name'
          sheet.cell(row=1,column=4).value='sub no'
          sheet.cell(row=1,column=3).value='schedule_sem'
          sheet.cell(row=1,column=2).value='register_sem'
          sheet.cell(row=1,column=1).value='roll no' 
          a=1
          for row1 in Read1:
                b=0
                c=0
                check=check+1
                if check==0:
                    continue
                course_master_file=open('course_master_dont_open_in_excel.csv','r')
                with course_master_file:
                    Read2=csv.reader(course_master_file)
                    for row2 in Read2:
                        #print(row1[3])

                        if row1[3]==row2[0]:
                            Store_LTP=row2[2]
                            Split_LTP=Store_LTP.split('-')
                            
                            #print(Separate_ltp)
                            if int(Split_LTP[0])>0:
                                b=b+1
                            if int(Split_LTP[1])>0:
                                b=b+1
                            if int(Split_LTP[2])>0:
                                b=b+1
                            #print(b)
                    feedback_submitted_file=open('course_feedback_submitted_by_students.csv','r')
                    with feedback_submitted_file:
                        Read3=csv.reader(feedback_submitted_file)
                        for row3 in Read3:
                          if row1[3]==row3[4] and row1[0]==row3[3] :
                              c=c+1
                              #print(c)
                        if c!=b:
                          sheet.cell(row=a,column=4).value=row1[3]
                          sheet.cell(row=a,column=3).value=row1[2]
                          sheet.cell(row=a,column=2).value=row1[1]
                          sheet.cell(row=a,column=1).value=row1[0]
                        #opening file studentinfo.csv
                        
                          with open('studentinfo.csv','r') as Student_info:
                              Read4=csv.reader(Student_info)
                              for row4 in Read4:
                                  if row1[0]==row4[1]:
                                      sheet.cell(row=a,column=8).value=row4[10]
                                      sheet.cell(row=a,column=7).value=row4[9]
                                      sheet.cell(row=a,column=6).value=row4[8]
                                      sheet.cell(row=a,column=5).value=row4[0] 
                          a=a+1
                          #print(a)
          W.save('course_feedback_remaining1.xlsx')     

                                 

feedback_not_submitted()