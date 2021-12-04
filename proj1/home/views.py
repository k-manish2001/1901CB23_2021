from django.shortcuts import render,HttpResponse
from openpyxl.styles import colors
from home.models import Index
from openpyxl.drawing.image import  Image
from openpyxl import Workbook, workbook
from openpyxl.styles import Font,DEFAULT_FONT
from openpyxl.styles.colors import Color
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
import csv
import os
import pandas as pd
DEFAULT_FONT.name = 'Times'
DEFAULT_FONT.size = 13

# Create your views here.
def index(request):
    if request.method=='POST' and 'log' in request.POST :
        c=request.POST.get('corr')
        n=request.POST.get('neg')
        f1=request.FILES["file1"]
        f2=request.FILES["file2"]
        y=pd.read_csv(f2)
        k=1
        check=0
        for data in y.itertuples():
            if data[7]=='ANSWER':
                check+=1
                break
        if check==0:
            return HttpResponse("response.csv file has no roll no. with named ANSWER")
     
        for data in y.itertuples():
            # k+=1
            # if k>10:
            #     break
                    
            if data[7]=='ANSWER':
                 stud_NR={}
                 num=28
                 for i in range(num):
                      stud_NR[i] = data[i+8]
                 op=Workbook()
                 op_sheet=op.active
                 iit_logo=Image("image/proj1.jpeg")
                 iit_logo.height=70
                 iit_logo.width=50
                 iit_logo.anchor='A1'
                 op_sheet.add_image(iit_logo)
                 op_sheet.cell(row=6 ,column=1).value ="Name:"
                 cell=op_sheet['B6']
                 cell.font=cell.font.copy(bold=True)
                 op_sheet.cell(row=6 ,column=2).value =data[4]
                 fontstyle=Font(size="18")
                 op_sheet.cell(row=5,column=3,value='Mark Sheet').font=fontstyle
                 op_sheet['C5'].font.copy(bold=True)
                 op_sheet.cell(row=6 ,column=4).value ="Exam:"
                 op_sheet.cell(row=6 ,column=5).value = "quiz"
                 op_sheet.cell(row=7 ,column=1).value = "Roll Number:"
                 op_sheet.cell(row=7 ,column=2).value = data[7]
                 op_sheet.cell(row=9 ,column=2).value = "Right" 
                 op_sheet.cell(row=9 ,column=3).value = "Wrong"
                 op_sheet.cell(row=9 ,column=4).value = "Not Attempt"
                 op_sheet.cell(row=9 ,column=5).value = "Max"
                 op_sheet.cell(row=10 ,column=1).value = "No."
                 op_sheet.cell(row=11,column=1).value = "Marking"
                 op_sheet.cell(row=11,column=2).value = float(c)
                 op_sheet.cell(row=11,column=3).value = float(n)
                 op_sheet.cell(row=12,column=1).value = "Total"
                 op_sheet.cell(row=15 ,column=1).value = "Student Ans"
                 op_sheet.cell(row=15 ,column=2).value = "Correct Ans"
                 op_sheet.cell(row=15 ,column=4).value = "Student Ans"
                 op_sheet.cell(row=15 ,column=5).value = "Correct Ans"
                 cell1=op_sheet['B7']
                 cell1.font=cell.font.copy(bold=True)
                 cell2=op_sheet['B9']
                 cell2.font=cell.font.copy(bold=True) 
                 cell3=op_sheet['E6']
                 cell3.font=cell.font.copy(bold=True)
                 cell4=op_sheet['C9']
                 cell4.font=cell.font.copy(bold=True)
                 cell5=op_sheet['D9']
                 cell5.font=cell.font.copy(bold=True)
                 cell6=op_sheet['A10']
                 cell6.font=cell.font.copy(bold=True)
                 cell7=op_sheet['A11']                
                 cell7.font=cell.font.copy(bold=True)
                 cell8=op_sheet['A12']
                 cell8.font=cell.font.copy(bold=True)
                 cell9=op_sheet['A15']
                 cell9.font=cell.font.copy(bold=True)
                 cell10=op_sheet['B15']
                 cell10.font=cell.font.copy(bold=True)
                 cell11=op_sheet['D15']
                 cell11.font=cell.font.copy(bold=True)
                 cell12=op_sheet['E15']
                 cell12.font=cell.font.copy(bold=True)
                 cell13=op_sheet['E9']
                 cell13.font=cell.font.copy(bold=True)
                 to=25
                 na=0
                 po=3
                 right=0
                 wrong=0
                 for i in range(to):

                     if str(data[i+8])=="nan":
                         na+=1
                     else:
                         if data[i+8]==stud_NR[i]:
                             right+=1
                             op_sheet.cell(row=i+16 ,column=1).value = data[i+8]
                             op_sheet[r'A'+str(16+i)].font=Font(color="FF00FF00")   
                         else:
                             wrong+=1
                             op_sheet.cell(row=i+16 ,column=1).value = data[i+8]
                             op_sheet[r'A'+str(16+i)].font=Font(color="FFFF0000")    
                     op_sheet.cell(row=i+16 ,column=2).value = stud_NR[i]
                     temp=i+16
                     op_sheet[r'B'+str(temp)].font=Font(color="FF0000FF")
                 for p in range(po):
                     if str(data[-3+p])=="nan":
                         na+=1
                     else:
                         if data[-3+p]==stud_NR[25+p]:
                             right+=1
                             op_sheet.cell(row=16+p ,column=4).value = data[-3+p]
                             op_sheet[r'D'+str(16+p)].font=Font(color="FF00FF00")

                         else:
                             wrong+=1 
                             op_sheet.cell(row=16+p ,column=4).value = data[-3+p]
                             op_sheet[r'D'+str(16+p)].font=Font(color="FFFF0000")       
                     op_sheet.cell(row=16+p ,column=5).value = stud_NR[25+p]
                     op_sheet[r'E'+str(16+p)].font=Font(color="FF0000FF")


                 op_sheet.cell(row=10 ,column=2).value = right
                 op_sheet.cell(row=10 ,column=3).value = wrong
                 op_sheet.cell(row=10,column=4).value = na
                 op_sheet.cell(row=10 ,column=5).value = right+wrong+na
                 op_sheet.cell(row=11 ,column=4).value = 0
                 op_sheet.cell(row=12,column=2).value = right*float(c)
                 op_sheet.cell(row=12,column=3).value = wrong*float(n)
                 op_sheet.cell(row=12,column=5).value = right*float(c)+wrong*float(n)
                 op_sheet['B10'].font=Font(color="FF00FF00")
                 op_sheet['B11'].font=Font(color="FF00FF00")
                 op_sheet['B12'].font=Font(color="FF00FF00")
                 op_sheet['E12'].font=Font(color="FF0000FF")
                 op_sheet['C10'].font=Font(color="FFFF0000")
                 op_sheet['C11'].font=Font(color="FFFF0000")
                 op_sheet['C12'].font=Font(color="FFFF0000") 

                # op_sheet.cell(row=16 ,column=5).value = stud_NR[25] 
                # op_sheet.cell(row=17 ,column=5).value = stud_NR[26]
                # op_sheet.cell(row=18 ,column=5).value = stud_NR[27]           
                 op.save(filename=r'sample_output/marksheet/'+data[7]+'.xlsx')
                                      
            elif data[7]=='Roll Number':
                 continue        
            else:

                op=Workbook()
                op_sheet=op.active
                iit_logo=Image("image/proj1.jpeg")
                iit_logo.height=60
                iit_logo.anchor='A1'
                op_sheet.add_image(iit_logo)
                op_sheet.cell(row=6 ,column=1).value ="Name:"
                cell=op_sheet['B6']
                cell.font=cell.font.copy(bold=True)
                op_sheet.cell(row=6 ,column=2).value =data[4]
                fontstyle=Font(size="18")
                op_sheet.cell(row=5,column=3,value='Mark Sheet').font=fontstyle
                op_sheet['C5'].font.copy(bold=True)
                op_sheet.cell(row=6 ,column=4).value ="Exam:"
                op_sheet.cell(row=6 ,column=5).value = "quiz"
                op_sheet.cell(row=7 ,column=1).value = "Roll Number:"
                op_sheet.cell(row=7 ,column=2).value = data[7]
                op_sheet.cell(row=9 ,column=2).value = "Right" 
                op_sheet.cell(row=9 ,column=3).value = "Wrong"
                op_sheet.cell(row=9 ,column=4).value = "Not Attempt"
                op_sheet.cell(row=9 ,column=5).value = "Max"
                op_sheet.cell(row=10 ,column=1).value = "No."
                op_sheet.cell(row=11,column=1).value = "Marking"
                op_sheet.cell(row=11,column=2).value = float(c)
                op_sheet.cell(row=11,column=3).value = float(n)
                op_sheet.cell(row=12,column=1).value = "Total"
                op_sheet.cell(row=15 ,column=1).value = "Student Ans"
                op_sheet.cell(row=15 ,column=2).value = "Correct Ans"
                op_sheet.cell(row=15 ,column=4).value = "Student Ans"
                op_sheet.cell(row=15 ,column=5).value = "Correct Ans"
                cell1=op_sheet['B7']
                cell1.font=cell.font.copy(bold=True)
                cell2=op_sheet['B9']
                cell2.font=cell.font.copy(bold=True) 
                cell3=op_sheet['E6']
                cell3.font=cell.font.copy(bold=True)
                cell4=op_sheet['C9']
                cell4.font=cell.font.copy(bold=True)
                cell5=op_sheet['D9']
                cell5.font=cell.font.copy(bold=True)
                cell6=op_sheet['A10']
                cell6.font=cell.font.copy(bold=True)
                cell7=op_sheet['A11']                
                cell7.font=cell.font.copy(bold=True)
                cell8=op_sheet['A12']
                cell8.font=cell.font.copy(bold=True)
                cell9=op_sheet['A15']
                cell9.font=cell.font.copy(bold=True)
                cell10=op_sheet['B15']
                cell10.font=cell.font.copy(bold=True)
                cell11=op_sheet['D15']
                cell11.font=cell.font.copy(bold=True)
                cell12=op_sheet['E15']
                cell12.font=cell.font.copy(bold=True)
                cell13=op_sheet['E9']
                cell13.font=cell.font.copy(bold=True)
                to=25
                na=0
                po=3
                right=0
                wrong=0
                for i in range(to):
                    if str(data[i+8])=="nan":
                        na+=1
                    else:
                        if data[i+8]==stud_NR[i]:
                            right+=1
                            op_sheet.cell(row=i+16 ,column=1).value = data[i+8]
                            op_sheet[r'A'+str(16+i)].font=Font(color="FF00FF00")   
                        else:
                            wrong+=1
                            op_sheet.cell(row=i+16 ,column=1).value = data[i+8]
                            op_sheet[r'A'+str(16+i)].font=Font(color="FFFF0000")    
                    op_sheet.cell(row=i+16 ,column=2).value = stud_NR[i]
                    temp=i+16
                    op_sheet[r'B'+str(temp)].font=Font(color="FF0000FF")
                for p in range(po):
                    if str(data[-3+p])=="nan":
                        na+=1
                    else:
                        if data[-3+p]==stud_NR[25+p]:
                            right+=1
                            op_sheet.cell(row=16+p ,column=4).value = data[-3+p]
                            op_sheet[r'D'+str(16+p)].font=Font(color="FF00FF00")

                        else:
                            wrong+=1 
                            op_sheet.cell(row=16+p ,column=4).value = data[-3+p]
                            op_sheet[r'D'+str(16+p)].font=Font(color="FFFF0000")       
                    op_sheet.cell(row=16+p ,column=5).value = stud_NR[25+p]
                    op_sheet[r'E'+str(16+p)].font=Font(color="FF0000FF")


                op_sheet.cell(row=10 ,column=2).value = right
                op_sheet.cell(row=10 ,column=3).value = wrong
                op_sheet.cell(row=10,column=4).value = na
                op_sheet.cell(row=10 ,column=5).value = right+wrong+na
                op_sheet.cell(row=11 ,column=4).value = 0
                op_sheet.cell(row=12,column=2).value = right*float(c)
                op_sheet.cell(row=12,column=3).value = wrong*float(n)
                op_sheet.cell(row=12,column=5).value = right*float(c)+wrong*float(n)
                op_sheet['B10'].font=Font(color="FF00FF00")
                op_sheet['B11'].font=Font(color="FF00FF00")
                op_sheet['B12'].font=Font(color="FF00FF00")
                op_sheet['E12'].font=Font(color="FF0000FF")
                op_sheet['C10'].font=Font(color="FFFF0000")
                op_sheet['C11'].font=Font(color="FFFF0000")
                op_sheet['C12'].font=Font(color="FFFF0000") 

                # op_sheet.cell(row=16 ,column=5).value = stud_NR[25] 
                # op_sheet.cell(row=17 ,column=5).value = stud_NR[26]
                # op_sheet.cell(row=18 ,column=5).value = stud_NR[27]           
                op.save(filename=r'sample_output/marksheet/'+data[7]+'.xlsx') 
        y1=pd.read_csv(f1)        
        for info in y1.itertuples():
            if info[1]=="roll" or info[1]=="ANSWER":
                continue
            else:
                if os.path.exists(r"sample_output/marksheet/"+info[1]+".xlsx"):
                    continue
                else:
                    # print(info[1])
                    op_missed=Workbook()
                    op_sheet=op_missed.active
                    op_missed.save(r"sample_output/marksheet/"+info[1]+".xlsx")

    elif request.method=='POST' and 'logg' in request.POST:
        c=request.POST.get('corr')
        n=request.POST.get('neg')
        f1=request.FILES["file1"]
        f2=request.FILES["file2"]  

        y=pd.read_csv(f2)   
        op1=Workbook()
        op1_sheet=op1.active
        op1_sheet.cell(row=1,column=1).value="timestatmp"
        op1_sheet.cell(row=1,column=2).value="Email address"
        op1_sheet.cell(row=1,column=3).value="Google_Score"
        op1_sheet.cell(row=1,column=4).value="Name"
        op1_sheet.cell(row=1,column=5).value="IITP webmail"
        op1_sheet.cell(row=1,column=6).value="Phone (10 digit only)"
        op1_sheet.cell(row=1,column=7).value="Score_After_Negative"
        op1_sheet.cell(row=1,column=8).value="Roll Number"
        for i in range(28):
            op1_sheet.cell(row=1,column=i+9).value="Unnamed"
        op1_sheet.cell(row=1,column=37).value="statusAns"
        
        row_inserted=2
        
        for data in y.itertuples():
            op1_sheet.cell(row=row_inserted,column=1).value=data[1]
            op1_sheet.cell(row=row_inserted,column=2).value=data[2]
            op1_sheet.cell(row=row_inserted,column=3).value=data[3]
            op1_sheet.cell(row=row_inserted,column=4).value=data[4]
            op1_sheet.cell(row=row_inserted,column=5).value=data[5]
            op1_sheet.cell(row=row_inserted,column=6).value=data[6]
            op1_sheet.cell(row=row_inserted,column=8).value=data[7]
            p=28
            for po in range(28):
                op1_sheet.cell(row=row_inserted,column=9+po).value=data[8+po]

            if data[7]=='ANSWER':
                    stud_NR={}
                    num=28
                    for i in range(num):
                        stud_NR[i] = data[i+8]
            t=28   
            not_a=0
            correct=0
            incorrect=0 
            for i in range(t):
                if str(data[i+8])=="nan":
                    not_a+=1
                else:
                    if data[i+8]==stud_NR[i]:
                        correct+=1    
                    else:
                        incorrect+=1  
            marks=(correct*float(c))+(incorrect*float(n)) 
            final_marks=str(marks)+"/140"
            op1_sheet.cell(row=row_inserted,column=7).value=final_marks          
            status="["+str(correct)+","+str(incorrect)+","+str(not_a)+"]"           
            op1_sheet.cell(row=row_inserted,column=37).value=status
            row_inserted+=1
        op1.save(filename=r'sample_output/marksheet/'+"concise_marksheet"+'.xlsx')  
   
    elif request.method=='POST' and 'log1' in request.POST:
        c=request.POST.get('corr')
        n=request.POST.get('neg')
        f1=request.FILES["file1"]
        f2=request.FILES["file2"]
        y3=pd.read_csv(f2)
        fromaddr="ankityadav55810@gmail.com"
        password='*******'#for privacy i have not added pass
        server=smtplib.SMTP('smtp.gmail.com',587)
        server.connect("smtp.gmail.com",587)
        server.starttls()
        server.login(fromaddr,password)
        
        l=1
        for data in y3.itertuples():
            print(data[2])
            # l+=1
            # if l>3:
            #     break
            if data[7]=="Roll Number" or data[7]=="ANSWER":
                continue
            else:   
                # li=["yadav.abhi0210@gmail.com","akaparalta2001@gmail.com"]
                li=[data[2],data[5]]
                for i in range(2):
                    x=li[i]
                    msg=MIMEMultipart()
                    msg['From']=fromaddr
                    msg['To']=x
                    msg['Subject']='Marksheet'
                    filename=r"sample_output/marksheet/"+data[7]+".xlsx"
                    attachement=open(filename,'rb')
                    p=MIMEBase('application','octet-stream')
                    p.set_payload((attachement).read())
                    encoders.encode_base64(p)
                    p.add_header('Content-Disposition',"attachment; filename=%s" % filename)
                    msg.attach(p)
                    text=msg.as_string()
                    server.send_message(msg)
        server.quit() 
        return HttpResponse("Your Email has been sent")           

              
                                         
    
    return render(request,'index.html')
